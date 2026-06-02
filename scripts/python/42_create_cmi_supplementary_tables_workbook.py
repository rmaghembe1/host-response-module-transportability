#!/usr/bin/env python3

from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

tables = [
    ("S1_candidate_register", "results/supplementary_tables/Supplementary_Table_S1_external_projection_candidate_register.tsv"),
    ("S2_locked_genes", "results/supplementary_tables/Supplementary_Table_S2_GSE211567_locked_discovery_module_genes.tsv"),
    ("S3A_identifier_coverage", "results/supplementary_tables/Supplementary_Table_S3A_GSE73461_module_identifier_coverage.tsv"),
    ("S3B_matched_genes", "results/supplementary_tables/Supplementary_Table_S3B_GSE73461_matched_locked_module_genes.tsv"),
    ("S3C_missing_genes", "results/supplementary_tables/Supplementary_Table_S3C_GSE73461_missing_locked_module_genes.tsv"),
    ("S3D_probe_choices", "results/supplementary_tables/Supplementary_Table_S3D_GSE73461_gene_probe_choice_for_projection.tsv"),
    ("S4A_scores_long", "results/supplementary_tables/Supplementary_Table_S4A_GSE73461_fixed_module_scores_long.tsv"),
    ("S4B_scores_wide", "results/supplementary_tables/Supplementary_Table_S4B_GSE73461_fixed_module_scores_wide.tsv"),
    ("S5A_primary_tests", "results/supplementary_tables/Supplementary_Table_S5A_GSE73461_primary_projection_tests.tsv"),
    ("S5B_sensitivity", "results/supplementary_tables/Supplementary_Table_S5B_GSE73461_primary_only_zscore_sensitivity_tests.tsv"),
    ("S5C_summary", "results/supplementary_tables/Supplementary_Table_S5C_GSE73461_manuscript_projection_summary.tsv"),
]

out = Path("submission/cmi_supplementary_tables_S1_to_S5.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)

missing = [p for _, p in tables if not Path(p).exists()]
if missing:
    raise SystemExit("Missing supplementary TSV files:\n" + "\n".join(missing))

wb = Workbook()
default = wb.active
wb.remove(default)

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
thin = Side(style="thin", color="808080")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

index_rows = []

for sheet_name, path in tables:
    ws = wb.create_sheet(title=sheet_name[:31])

    with open(path, newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        rows = list(reader)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = wrap
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if rows:
        n_cols = len(rows[0])
        sample_rows = rows[:101]
        for c_idx in range(1, n_cols + 1):
            max_len = 12
            for row in sample_rows:
                if c_idx <= len(row):
                    max_len = max(max_len, len(str(row[c_idx - 1])) + 2)
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len, 45)

    index_rows.append([sheet_name, path, max(len(rows) - 1, 0), len(rows[0]) if rows else 0])

idx = wb.create_sheet(title="Index", index=0)
idx.append(["sheet", "source_tsv", "rows", "columns"])
for row in index_rows:
    idx.append(row)

for cell in idx[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = wrap

idx.freeze_panes = "A2"
idx.auto_filter.ref = idx.dimensions
for c_idx in range(1, 5):
    idx.column_dimensions[get_column_letter(c_idx)].width = 60 if c_idx == 2 else 24

wb.save(out)
print(f"Wrote: {out}")
