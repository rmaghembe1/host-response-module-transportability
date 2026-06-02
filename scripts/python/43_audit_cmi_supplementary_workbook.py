#!/usr/bin/env python3

from pathlib import Path
from openpyxl import load_workbook
import csv

path = Path("submission/cmi_supplementary_tables_S1_to_S5.xlsx")

expected_sheets = [
    "Index",
    "S1_candidate_register",
    "S2_locked_genes",
    "S3A_identifier_coverage",
    "S3B_matched_genes",
    "S3C_missing_genes",
    "S3D_probe_choices",
    "S4A_scores_long",
    "S4B_scores_wide",
    "S5A_primary_tests",
    "S5B_sensitivity",
    "S5C_summary",
]

if not path.exists():
    raise SystemExit(f"Missing workbook: {path}")

wb = load_workbook(path, read_only=True, data_only=True)
actual_sheets = wb.sheetnames

rows = []
for sheet in expected_sheets:
    present = sheet in actual_sheets
    if present:
        ws = wb[sheet]
        rows.append({
            "sheet": sheet,
            "present": "TRUE",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "status": "present",
        })
    else:
        rows.append({
            "sheet": sheet,
            "present": "FALSE",
            "max_row": "",
            "max_column": "",
            "status": "missing",
        })

extra = [s for s in actual_sheets if s not in expected_sheets]

out = Path("results/audits/cmi_supplementary_workbook_audit.tsv")
out.parent.mkdir(parents=True, exist_ok=True)

with out.open("w", newline="") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=["sheet", "present", "max_row", "max_column", "status"],
        delimiter="\t"
    )
    writer.writeheader()
    writer.writerows(rows)

present = sum(r["status"] == "present" for r in rows)
missing = sum(r["status"] == "missing" for r in rows)

summary = Path("results/audits/cmi_supplementary_workbook_audit_summary.md")
summary.write_text(
    "# CMI Supplementary Workbook Audit Summary\n\n"
    f"Workbook: `{path}`\n\n"
    f"Expected sheets: {len(expected_sheets)}\n\n"
    f"Present sheets: {present}\n\n"
    f"Missing sheets: {missing}\n\n"
    f"Extra sheets: {len(extra)}\n\n"
    f"Extra sheet names: {', '.join(extra) if extra else 'None'}\n\n"
    f"Detailed audit: `{out}`\n"
)

print(summary.read_text())

if missing:
    raise SystemExit("Missing expected supplementary workbook sheets.")
